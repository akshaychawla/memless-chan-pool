import numpy as np
import torch.nn as nn
import torch
import random
from PRCN import PRCNPTN, FastPRCNPTN 
import os, sys, time
from pprint import pprint
from openpyxl import Workbook, load_workbook
import argparse

def runTest(params):
    """
    Run performance test with given parameters for PRCNv1 and PRCNv2
    """
    ip_chans, op_chans = params["ip_chans"], params["op_chans"]
    h, w = params["h"], params["w"]
    batch_size = params["batch_size"]
    G, CMP = params["G"], params["CMP"]
    module = None
    if params["module"] == "PRCNPTN": 
        module = PRCNPTN
    elif params["module"] == "FastPRCNPTN": 
        module = FastPRCNPTN 
    else: 
        raise NotImplementedError("module {} cannot be tested".format(params["module"]))

    obj = module(
                ip_chans, op_chans,
                G=G, CMP=CMP, kernel_size=3, 
                padding=1, stride=1
            ) 
    obj = obj.cuda()
    obj.eval() 

    x = torch.randn(batch_size, ip_chans, h, w).float().cuda() 
    # Pre-cache run 
    for _ in range(5):
        with torch.no_grad():
            _ = obj(x) 
    # Timed run 
    elapsed = [] 
    for _ in range(params["numsamples"]): 
        with torch.no_grad(): 
            start = time.time()  
            _ = obj(x) 
            elapsed.append(time.time() - start)
    return float(np.mean(elapsed))

def cli(): 
    parser = argparse.ArgumentParser() 
    params = {
        "module": "PRCNPTN",
        "ip_chans":16, 
        "op_chans":64,
        "h": 128, "w": 128, 
        "batch_size":16, 
        "G": 12, "CMP":4,
        "numsamples": 50,
        "xlsx": "./performance.xlsx"
        } 
    for key,val in params.items():
        parser.add_argument("--"+key, type=type(val), default=val)
    args = parser.parse_args() 
    return vars(args)

if __name__ == "__main__":
    args = cli() 
    pprint(args)
    elapsed = runTest(args)
    print("mean time: {:.8f}".format(elapsed))
    mem = torch.cuda.max_memory_allocated(0)
    mem = mem / (1024.0*1024.0)
    print("max mem: {:.2f}".format(mem))

    # write to xlsx file 
    if os.path.exists(args["xlsx"]) == False:
        wb = Workbook()
    else: 
        wb = load_workbook(args["xlsx"])
    ws = wb.active 
    # find empty row 
    idx = 1 
    while ws["A"+str(idx)].value is not None:
        idx += 1 
    store_tuple = ( 
            time.asctime(time.gmtime(time.time())), args["module"], args["batch_size"], 
            args["ip_chans"], args["op_chans"], args["h"], args["w"], args["G"], 
            args["CMP"], mem, elapsed 
    )
    for col_idx, store_val in enumerate(store_tuple): 
        ws[str(chr(65+col_idx))+str(idx)] = str(store_val)
    wb.save(args["xlsx"])

        



